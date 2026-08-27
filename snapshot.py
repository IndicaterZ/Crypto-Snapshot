import requests
import os
import concurrent.futures
import hmac
import hashlib
import time
import uuid
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

EXCEL_FILE = "crypto_volume.xlsx"

def get_binance_th_data():
    try:
        exchange_info = requests.get('https://api.binance.th/api/v1/exchangeInfo')
        exchange_info.raise_for_status()
        symbols = [s['symbol'] for s in exchange_info.json()['symbols'] if s['symbol'].endswith('THB')]
        
        def fetch_ticker(symbol):
            try:
                resp = requests.get(f'https://api.binance.th/api/v1/ticker/24hr?symbol={symbol}')
                resp.raise_for_status()
                return resp.json()
            except Exception:
                return None
                
        tickers = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            results = executor.map(fetch_ticker, symbols)
            tickers = [t for t in results if t is not None]
        
        thb_pairs = []
        for t in tickers:
            try:
                t['quoteVolume'] = float(t['quoteVolume'])
                t['lastPrice'] = float(t.get('lastPrice', 0))
                thb_pairs.append(t)
            except (ValueError, KeyError):
                continue
        
        thb_pairs.sort(key=lambda x: x['quoteVolume'], reverse=True)
        return thb_pairs
    except Exception as e:
        print(f"Error fetching Binance TH data: {e}")
        return []

def get_bitkub_data():
    try:
        response = requests.get('https://api.bitkub.com/api/market/ticker')
        response.raise_for_status()
        data = response.json()
        
        thb_pairs = []
        for symbol, info in data.items():
            if symbol.startswith('THB_'):
                try:
                    quote_vol = float(info['quoteVolume'])
                    last_price = float(info['last'])
                    thb_pairs.append({
                        'symbol': symbol,
                        'quoteVolume': quote_vol,
                        'lastPrice': last_price
                    })
                except (ValueError, KeyError):
                    continue
                    
        thb_pairs.sort(key=lambda x: x['quoteVolume'], reverse=True)
        return thb_pairs
    except Exception as e:
        print(f"Error fetching Bitkub data: {e}")
        return []

def get_orbix_data():
    try:
        response = requests.get('https://satangcorp.com/api/v3/ticker/24hr')
        response.raise_for_status()
        data = response.json()
        
        thb_pairs = []
        for d in data:
            if d['symbol'].endswith('_thb'):
                try:
                    quote_vol = float(d['quoteVolume'])
                    last_price = float(d['lastPrice'])
                    thb_pairs.append({
                        'symbol': d['symbol'].upper(),
                        'quoteVolume': quote_vol,
                        'lastPrice': last_price
                    })
                except (ValueError, KeyError):
                    continue
        
        thb_pairs.sort(key=lambda x: x['quoteVolume'], reverse=True)
        return thb_pairs
    except Exception as e:
        print(f"Error fetching Orbix data: {e}")
        return []

def get_innovestx_data():
    api_key = os.environ.get('INNVX_API_KEY')
    api_secret = os.environ.get('INNVX_API_SECRET')
    
    if not api_key or not api_secret:
        print("InnovestX API keys not found in environment variables.")
        return []
        
    try:
        timestamp = str(int(time.time() * 1000))
        uid = str(uuid.uuid4())
        method = "GET"
        path = "/api/v1/digital-asset/symbols" # Endpoint might need to be adjusted based on actual docs
        host = "api.innovestxonline.com"
        query = ""
        content_type = ""
        body = ""
        
        content_to_sign = f"{api_key}{method}{host}{path}{query}{content_type}{uid}{timestamp}{body}"
        signature = hmac.new(
            api_secret.encode('utf-8'),
            content_to_sign.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            'X-INVX-APIKEY': api_key,
            'X-INVX-SIGNATURE': signature,
            'X-INVX-TIMESTAMP': timestamp,
            'X-INVX-REQUEST-UID': uid,
            'Accept': 'application/json'
        }
        
        response = requests.get(f'https://{host}{path}', headers=headers)
        response.raise_for_status()
        data = response.json()
        
        thb_pairs = []
        items = data.get('data', data) if isinstance(data, dict) else data
        for d in items:
            symbol = str(d.get('symbol', ''))
            if symbol.endswith('THB') or symbol.startswith('THB_'):
                try:
                    quote_vol = float(d.get('quoteVolume', 0))
                    last_price = float(d.get('lastPrice', 0))
                    thb_pairs.append({
                        'symbol': symbol,
                        'quoteVolume': quote_vol,
                        'lastPrice': last_price
                    })
                except (ValueError, KeyError):
                    continue
                    
        thb_pairs.sort(key=lambda x: x['quoteVolume'], reverse=True)
        return thb_pairs
    except Exception as e:
        print(f"Error fetching InnovestX data: {e}")
        return []

def update_excel(bitkub_data, binance_data, orbix_data, innovestx_data):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Volume Snapshot"
    
    # 1. Update Top Section (Rows 1-6) - Only show Top 5
    for row in range(1, 7):
        for col in range(1, 15):
            ws.cell(row=row, column=col).value = None

    headers = [
        ("Bitkub", 2), ("Ave. daily vol.", 3),
        ("Binance TH", 5), ("Ave. daily vol.", 6),
        ("Orbix", 8), ("Ave. daily vol.", 9),
        ("InnovestX", 11), ("Ave. daily vol.", 12)
    ]
    for text, col in headers:
        ws.cell(row=1, column=col).value = text

    groups = [
        (bitkub_data, 2),
        (binance_data, 5),
        (orbix_data, 8),
        (innovestx_data, 11)
    ]

    for i in range(5):
        row = i + 2
        ws.cell(row=row, column=1).value = i + 1
        
        for data, start_col in groups:
            if i < len(data):
                ws.cell(row=row, column=start_col).value = data[i]['symbol']
                c = ws.cell(row=row, column=start_col+1)
                c.value = data[i]['quoteVolume']
                c.number_format = '#,##0.00'

    # 2. Historical Section (Record ALL coins)
    ws.cell(row=13, column=1).value = "Date"
    
    header_map = {}
    group_ends = [1, 1, 1, 1] # Bitkub, Binance, Orbix, InnovestX
    
    # Read existing map
    for col in range(2, 2000):
        val = ws.cell(row=13, column=col).value
        if val and not str(val).endswith("7d-Avg"):
            header_map[val] = col
            
    # Assign existing columns to groups to find their boundaries
    for val, col in header_map.items():
        if val.startswith("THB_"): group_ends[0] = max(group_ends[0], col)
        elif "_" not in val and val.endswith("THB"): group_ends[1] = max(group_ends[1], col)
        elif val.endswith("_THB"): group_ends[2] = max(group_ends[2], col)
        else: group_ends[3] = max(group_ends[3], col) # Fallback to InnovestX
        
    def add_missing_symbols(data, group_idx):
        nonlocal header_map, group_ends
        for d in data:
            sym = d['symbol']
            if sym not in header_map:
                insert_col = group_ends[group_idx]
                if insert_col == 1:
                    prev_end = 1
                    for i in range(group_idx - 1, -1, -1):
                        if group_ends[i] > 1:
                            prev_end = group_ends[i]
                            break
                    insert_col = prev_end + 1
                else:
                    insert_col += 1
                
                needs_shift = False
                for i in range(group_idx + 1, 4):
                    if group_ends[i] >= insert_col:
                        needs_shift = True
                        break
                        
                if needs_shift:
                    ws.insert_cols(insert_col)
                    for i in range(group_idx + 1, 4):
                        if group_ends[i] > 1:
                            group_ends[i] += 1
                            
                ws.cell(row=13, column=insert_col).value = sym
                header_map[sym] = insert_col
                if group_ends[group_idx] == 1: group_ends[group_idx] = insert_col
                else: group_ends[group_idx] = max(group_ends[group_idx], insert_col)
                
                if "BTC" in sym:
                    insert_col += 1
                    if needs_shift:
                        ws.insert_cols(insert_col)
                        for i in range(group_idx + 1, 4):
                            if group_ends[i] > 1:
                                group_ends[i] += 1
                    ws.cell(row=13, column=insert_col).value = f"{sym} 7d-Avg"
                    group_ends[group_idx] = max(group_ends[group_idx], insert_col)
                    
                header_map.clear()
                for c in range(2, 2000):
                    v = ws.cell(row=13, column=c).value
                    if v and not str(v).endswith("7d-Avg"):
                        header_map[v] = c

    add_missing_symbols(bitkub_data, 0)
    add_missing_symbols(binance_data, 1)
    add_missing_symbols(orbix_data, 2)
    add_missing_symbols(innovestx_data, 3)

    # Find next empty row for historical data
    next_row = 14
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Write Date
    th_tz = timezone(timedelta(hours=7))
    current_date = datetime.now(th_tz).day
    ws.cell(row=next_row, column=1).value = current_date

    # Write volumes and formulas
    all_coins = []
    for grp in groups:
        all_coins.extend([(d['symbol'], d['quoteVolume']) for d in grp[0]])
        
    for symbol, vol in all_coins:
        vol_col = header_map.get(symbol)
        if not vol_col: continue
        
        c_vol = ws.cell(row=next_row, column=vol_col)
        c_vol.value = vol
        c_vol.number_format = '#,##0.00'
        
        if "BTC" in symbol:
            if next_row >= 20:
                col_letter = get_column_letter(vol_col)
                formula = f"=AVERAGE({col_letter}{next_row-6}:{col_letter}{next_row})"
                c_avg = ws.cell(row=next_row, column=vol_col+1)
                c_avg.value = formula
                c_avg.number_format = '#,##0.00'

    wb.save(EXCEL_FILE)
    print(f"Successfully updated {EXCEL_FILE} with data for all coins.")

def main():
    print("Starting snapshot for ALL coins...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        f_binance = executor.submit(get_binance_th_data)
        f_bitkub = executor.submit(get_bitkub_data)
        f_orbix = executor.submit(get_orbix_data)
        f_innovestx = executor.submit(get_innovestx_data)
        
        binance_data = f_binance.result()
        bitkub_data = f_bitkub.result()
        orbix_data = f_orbix.result()
        innovestx_data = f_innovestx.result()
    
    if any([binance_data, bitkub_data, orbix_data, innovestx_data]):
        update_excel(bitkub_data, binance_data, orbix_data, innovestx_data)
    else:
        print("No data fetched from any exchange.")

if __name__ == "__main__":
    main()
